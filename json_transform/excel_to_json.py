import json
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / 'input'
OUTPUT_DIR = BASE_DIR / 'output'

ruta_excel = INPUT_DIR / 'data.xlsx'

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def nombre_archivo_seguro(nombre_hoja: str) -> str:
    """Convierte el nombre de la hoja en un nombre de archivo válido."""
    nombre = nombre_hoja.strip() or 'hoja'
    nombre = re.sub(r'[<>:"/\\|?*]', '_', nombre)
    return f'{nombre}.json'


def celda_vacia(valor) -> bool:
    return valor is None or (isinstance(valor, str) and valor.strip() == '')


def fila_vacia(fila) -> bool:
    return all(celda_vacia(v) for v in fila)


def expandir_celdas_combinadas(ws) -> None:
    """Copia el valor de cada celda combinada a todas las celdas del rango."""
    for merge in list(ws.merged_cells.ranges):
        valor = ws.cell(merge.min_row, merge.min_col).value
        ws.unmerge_cells(str(merge))
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                ws.cell(row, col).value = valor


def normalizar_cabeceras(cabeceras: list) -> list[str]:
    """Nombra cabeceras vacías y evita duplicados (Grupo, Grupo.1, ...)."""
    vistas: dict[str, int] = {}
    resultado = []
    for i, raw in enumerate(cabeceras):
        nombre = str(raw).strip() if not celda_vacia(raw) else f'Unnamed: {i}'
        if nombre in vistas:
            vistas[nombre] += 1
            nombre = f'{nombre}.{vistas[nombre]}'
        else:
            vistas[nombre] = 0
        resultado.append(nombre)
    return resultado


def dividir_en_bloques(filas: list[list]) -> list[list[list]]:
    """Separa la hoja en bloques usando filas completamente vacías."""
    bloques = []
    actual = []
    for fila in filas:
        if fila_vacia(fila):
            if actual:
                bloques.append(actual)
                actual = []
            continue
        actual.append(fila)
    if actual:
        bloques.append(actual)
    return bloques


def puntuar_bloque(bloque: list[list]) -> tuple:
    """Prioriza el bloque con más cabeceras útiles y más datos."""
    if len(bloque) < 2:
        return (0, 0, 0)
    cabeceras = bloque[0]
    cabeceras_llenas = sum(1 for c in cabeceras if not celda_vacia(c))
    celdas_datos = sum(1 for fila in bloque[1:] for c in fila if not celda_vacia(c))
    return (cabeceras_llenas, celdas_datos, len(bloque))


def bloque_a_registros(bloque: list[list]) -> list[dict]:
    # Quitar columnas totalmente vacías (cabecera + datos)
    n_cols = max(len(f) for f in bloque)
    filas = [f + [None] * (n_cols - len(f)) for f in bloque]
    cols_utiles = [
        i for i in range(n_cols)
        if any(not celda_vacia(fila[i]) for fila in filas)
    ]
    filas = [[fila[i] for i in cols_utiles] for fila in filas]

    cabeceras = normalizar_cabeceras(filas[0])
    registros = []
    for fila in filas[1:]:
        if fila_vacia(fila):
            continue
        registros.append({cabeceras[i]: fila[i] for i in range(len(cabeceras))})
    return registros


def hoja_a_registros(ws) -> list[dict]:
    expandir_celdas_combinadas(ws)

    filas = [list(row) for row in ws.iter_rows(values_only=True)]
    # Recortar filas vacías al final
    while filas and fila_vacia(filas[-1]):
        filas.pop()
    if not filas:
        return []

    bloques = dividir_en_bloques(filas)
    if not bloques:
        return []

    bloque = max(bloques, key=puntuar_bloque)
    return bloque_a_registros(bloque)


wb = openpyxl.load_workbook(ruta_excel)
hojas_visibles = [ws for ws in wb.worksheets if ws.sheet_state == 'visible']

for ws in hojas_visibles:
    data = hoja_a_registros(ws)
    ruta_json = OUTPUT_DIR / nombre_archivo_seguro(ws.title)

    with open(ruta_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    print(f'Archivo JSON generado: {ruta_json} ({len(data)} registros)')

wb.close()
print(f'Total de hojas convertidas: {len(hojas_visibles)}')
